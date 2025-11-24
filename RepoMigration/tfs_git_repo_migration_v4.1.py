from csv import excel
import os
import subprocess
import requests
import json
import logging
from datetime import datetime
import base64
import time
from urllib.parse import quote, quote_plus
import argparse
import pandas as pd
import sys
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from github import Github, GithubException
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
import shutil
import stat
from requests.auth import HTTPBasicAuth
import nacl.encoding
import nacl.public 

# --- Ensure folders exist ---
log_dir = os.path.join(os.getcwd(), "tfs_repo_migration", "tfs_git_migration_log")
os.makedirs(log_dir, exist_ok=True)

# --- Timestamped log file path ---
timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
log_file = os.path.join(log_dir, f"script_log_{timestamp}.log")

# --- Logging Setup ---
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.FileHandler(log_file, mode='w', encoding='utf-8'),
        logging.StreamHandler()
    ]
)

logging.info(f"Logging started. Log file at: {log_file}")

def find_bfg_jar():
    operation_dir = os.getcwd()  
    for root, dirs, files in os.walk(operation_dir):
        for file in files:
            if file.endswith(".jar") and "bfg" in file.lower():  
                return os.path.join(root, file)
    raise FileNotFoundError("BFG JAR file not found in the operation directory.")

try:
    BFG_JAR_PATH = find_bfg_jar()
    logging.info(f"BFG JAR file found at: {BFG_JAR_PATH}")
except FileNotFoundError as e:
    logging.error(str(e))
    sys.exit(1)
    
# PROXIES
PROXIES = {
    "http": "http://your.proxy.server:port",
    "https": "http://your.proxy.server:port"
}

TOKEN_PATHS = {
    "github": os.path.expanduser("~/.github_token"),
    "tfs": os.path.expanduser("~/.tfs_token")
}

CHECKPOINT_FILE = os.path.join(os.getcwd(), "tfs_git_checkpoint.json")
migration_summary = {
    "migrated_repos": [],
    "skipped_repos": [],
    "failed_repos": []
}

def encode_path_segment(str_to_encode):
    """
    Percent-encode a single path segment for URLs.
    Use safe='' to ensure spaces become %20 (not '+').
    """
    if str_to_encode is None:
        return ""
    return quote(str(str_to_encode), safe='')


def get_tfs_domain_from_template(template_file):
    if not os.path.exists(template_file):
        raise FileNotFoundError(f"Template config file not found: {template_file}")

    with open(template_file, "r") as f:
        config = json.load(f)

    tfs_url = config.get("tfs_url", "").strip()
    return tfs_url

# or however you load it
def validate_token(service, token, tfs_domain):
    if service == "github":
        headers = {"Authorization": f"token {token}"}
        response = requests.get("https://api.github.com/user", headers=headers)
        return response.status_code == 200
    elif service == "tfs":
        headers = {"Authorization": f"Basic {base64.b64encode(f':{token}'.encode()).decode()}"}
        url = f"http://{tfs_domain}"
        print(url)
        try:
            response = requests.get(url, headers=headers)
            return response.status_code == 200
        except requests.RequestException:
            return False
    return False

def load_tokens(token_file):
    """
    Loads GitHub and TFS tokens from a required JSON file.
    """
    if not token_file or not os.path.exists(token_file):
        logging.error(f"Token file not found or not provided: {token_file}")
        sys.exit(1)
 
    try:
        with open(token_file, "r") as f:
            token_data = json.load(f)
    except Exception as e:
        logging.error(f"Failed to read token file '{token_file}': {e}")
        sys.exit(1)
 
    # Ensure both tokens are present
    for service in ["github", "tfs"]:
        token = token_data.get(service)
        if not token:
            logging.error(f"{service.upper()} token is missing in the token file.")
            sys.exit(1)
        if not validate_token(service, token, tfs_domain):
            logging.error(f"{service.upper()} token is invalid.")
            sys.exit(1)
 
    logging.info(f"Successfully loaded and validated tokens from {token_file}")
    return token_data

def save_summary_log(output_path="migration_summary_log.json"):
    with open(output_path, "w") as file:
        json.dump(migration_summary, file, indent=4)
    logging.info(f"Migration summary saved to: {output_path}")

def generate_config_jsons(excel_file, template_file, output_folder="generated_config_json"):
    os.makedirs(output_folder, exist_ok=True)
    with open(template_file, 'r') as f:
        template = json.load(f)

    global github_pat, tfs_pat
    azure_pat = tfs_pat
    github_token = github_pat
    df = pd.read_excel(excel_file, sheet_name="migration")

    generated_files = []
    row_mapping = {}

    for idx, row in df.iterrows():
        ignore_val = str(row.get("IGNORE", "")).strip().lower()
        if ignore_val == "yes":
            continue  # Skip if IGNORE = "yes"

        config = template.copy()
        config["azure_devops_pat_token"] = azure_pat
        config["github_token"] = github_token
        config["azure_devops_organization"] = row.get("COLLECTION_NAME", "")
        config["azure_devops_project"] = row.get("PROJECT_NAME", "")
        config["tfs_source_repo"] = row.get("TFS_REPO", "")
        config["github_organization"] = row.get("GITHUB_ORGANIZATION", "")
        config["github_target_repo"] = row.get("GITHUB_REPO", "")
        config["has_large_files"] = row.get("LARGE FILES", "no")  # Check for "LARGE FILES" column
        branches_cell = row.get("BRANCH_LIST", "all")
        if isinstance(branches_cell, str) and branches_cell.strip().lower() != "all":
            config["specific_branches"] = [branch.strip() for branch in branches_cell.split(",")]
        else:
            config["specific_branches"] = "all"

        filename = f"generated_config_{config['github_target_repo']}.json"
        filepath = os.path.join(output_folder, filename)
        with open(filepath, 'w') as f:
            json.dump(config, f, indent=4)
        generated_files.append(filepath)
        row_mapping[filepath] = idx  # Store Excel row index

    return generated_files, df, row_mapping

def setup_logging(log_dir):
    log_filename = datetime.now().strftime("script_log_%Y-%m-%d_%H-%M-%S.log")
    log_path = os.path.join(log_dir, log_filename)
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s [%(levelname)s]: %(message)s",
        handlers=[
            logging.FileHandler(log_path, mode='w'),
            logging.StreamHandler()
        ]
    )
    logging.info(f"Logging initialized: {log_path}")

def load_config_from_path(config_path):
    with open(config_path, "r") as config_file:
        return json.load(config_file)

def get_auth_header(pat):
    auth_str = f":{pat}".encode("utf-8")
    auth_b64 = base64.b64encode(auth_str).decode("utf-8")
    return {"Authorization": f"Basic {auth_b64}"}

def get_github_auth_header(token):
    return {"Authorization": f"token {token}"}

def save_checkpoint(collection_name, project_name, repo_name):
    checkpoint = {}
    if os.path.exists(CHECKPOINT_FILE):
        with open(CHECKPOINT_FILE, "r") as file:
            checkpoint = json.load(file)
    if collection_name not in checkpoint:
        checkpoint[collection_name] = {}
    if project_name not in checkpoint[collection_name]:
        checkpoint[collection_name][project_name] = []
    if repo_name not in checkpoint[collection_name][project_name]:
        checkpoint[collection_name][project_name].append(repo_name)
    with open(CHECKPOINT_FILE, "w") as file:
        json.dump(checkpoint, file, indent=4)
    logging.info(f"Checkpoint updated: {repo_name} saved under {collection_name}/{project_name}.")

def load_checkpoint():
    if not os.path.exists(CHECKPOINT_FILE):
        with open(CHECKPOINT_FILE, "w") as file:
            json.dump({}, file)
    with open(CHECKPOINT_FILE, "r") as file:
        return json.load(file)

def retry_subprocess(command, retries=3, enable_proxy=False):
    last_error = ""
    for attempt in range(retries):
        try:
            env = os.environ.copy()
            if enable_proxy:
                env['https_proxy'] = PROXIES["https"]
                env['http_proxy'] = PROXIES["http"]

            result = subprocess.run(
                command,
                check=True,
                env=env,
                stdout=subprocess.PIPE,
                stderr=subprocess.PIPE,
                text=True
            )

            logging.info(f"Command output:\n{result.stdout}")
            if result.stderr:
                logging.warning(f"Command stderr:\n{result.stderr}")

            return result.stdout

        except subprocess.CalledProcessError as e:
            last_error = e.stderr.strip()
            logging.error(f"Attempt {attempt + 1} failed: {last_error}")
            time.sleep(2 ** attempt)

    raise Exception(f"Command failed after {retries} retries: {' '.join(command)}\nError: {last_error}")

def migrate_specific_branches(repo_dir, github_repo_url, branches, enable_proxy=False):
    for branch in branches:
        logging.info(f"Pushing specific branch: {branch}")
        retry_subprocess(["git", "fetch", "origin", branch])
        retry_subprocess(["git", "push", github_repo_url, f"refs/remotes/origin/{branch}:refs/heads/{branch}"], enable_proxy=enable_proxy)

def create_directory_structure(base_dir, org, project):
    org_dir = os.path.join(base_dir, org)
    project_dir = os.path.join(org_dir, project)
    os.makedirs(project_dir, exist_ok=True)
    return project_dir

def check_github_repo_exists(github_org, repo_name, github_token, enable_proxy=False):
    url = f"https://api.github.com/repos/{github_org}/{repo_name}"
    headers = {"Authorization": f"token {github_token}"}
    response = github_request("GET", url, headers=headers, enable_proxy=enable_proxy)
    return response and response.status_code == 200

def move_bfg_report_to_log(collection_name, project_name, repo_name):
    base_dir = os.path.join(os.getcwd(), "tfs_repo_migration")
    src = os.path.join(base_dir, "tfs_git_repo", collection_name, project_name, f"{repo_name}.bfg-report")
    dest_dir = os.path.join(base_dir, "tfs_git_migration_log", f"{collection_name}_{project_name}")
    dest = os.path.join(dest_dir, f"{repo_name}.bfg-report")

    try:
        if os.path.exists(src):
            os.makedirs(dest_dir, exist_ok=True)
            # If dest exists, remove it first to replace
            if os.path.isdir(dest):
                shutil.rmtree(dest)
            elif os.path.isfile(dest):
                os.remove(dest)
            shutil.move(src, dest)

            # Append timestamp to the moved item (preserves extension if present)
            ts = datetime.now().strftime("%Y%m%d_%H%M%S")
            base, ext = os.path.splitext(dest)
            dest_ts = f"{base}_{ts}{ext}"

            # If target with timestamp already exists, add a numeric suffix
            counter = 1
            while os.path.exists(dest_ts):
                dest_ts = f"{base}_{ts}_{counter}{ext}"
                counter += 1

            os.rename(dest, dest_ts)
            logging.info(f"Moved BFG report from {src} to {dest}")
        else:
            logging.info(f"No BFG report found at {src}")
    except Exception as e:
        logging.warning(f"Failed to move BFG report from {src} to {dest}: {e}")

def clone_and_push_repositories(config, only_clone_repos, enable_proxy=False):
    global github_pat, tfs_pat, migration_summary

    azure_pat = tfs_pat
    github_token = github_pat

    azure_org = str(config["azure_devops_organization"])
    azure_project = config["azure_devops_project"]
    tfs_repo = config["tfs_source_repo"]
    git_repo = config["github_target_repo"]
    tfs_domain = config["tfs_url"]
    collection_name = azure_org
    specific_branches = config.get("specific_branches", "all")
    has_large_files = str(config.get("has_large_files", "no")).strip().lower() == "yes"

    if not tfs_repo or not azure_project:
        raise ValueError("TFS_REPO or PROJECT_NAME is missing in config.")

    if not tfs_domain.startswith("http"):
        tfs_domain = "http://" + tfs_domain
    tfs_domain = tfs_domain.rstrip("/")

    base_dir = os.path.join(os.getcwd(), "tfs_repo_migration")
    repo_base_dir = os.path.join(base_dir, "tfs_git_repo")
    log_dir = os.path.join(base_dir, "tfs_git_migration_log")
    os.makedirs(repo_base_dir, exist_ok=True)
    os.makedirs(log_dir, exist_ok=True)

    project_dir = create_directory_structure(repo_base_dir, azure_org, azure_project)
    # Encode path segments for URLs only
    coll_enc = encode_path_segment(azure_org)
    proj_enc = encode_path_segment(azure_project)

    url = f"{tfs_domain}/{coll_enc}/{proj_enc}/_apis/git/repositories?api-version=5.0"
    headers = get_auth_header(azure_pat)
    response = requests.get(url, headers=headers)
    if response.status_code != 200:
        raise Exception(f"Failed to fetch repositories: {response.text}")

    repositories = response.json().get("value", [])
    repo_names = [repo["name"] for repo in repositories]

    if tfs_repo not in repo_names:
        raise Exception(f"Repository {tfs_repo} not found in TFS.")

    checkpoint = load_checkpoint()
    repo = next((r for r in repositories if r["name"] == tfs_repo), None)
    original_cwd = os.getcwd()

    try:
        if repo:
            repo_name = str(repo["name"])
            if (
                collection_name in checkpoint
                and azure_project in checkpoint[collection_name]
                and repo_name in checkpoint[collection_name][azure_project]
            ):
                logging.info(f"Skipping {repo_name}, already migrated.")
                migration_summary.setdefault("skipped_repos", []).append(repo_name)
                return

            repo_dir = os.path.join(project_dir, repo_name)
            # URL-encode the PAT if embedding it in the clone URL (still not recommended security-wise)
            azure_pat_enc = quote(str(azure_pat), safe='')
            repo_enc = encode_path_segment(repo_name)

            # Build the azure repo URL using encoded path segments
            azure_repo_url = f"{tfs_domain.replace('http://', f'http://{azure_pat_enc}@')}/{coll_enc}/{proj_enc}/_git/{repo_enc}"

            if not os.path.exists(repo_dir):
                if specific_branches == "all":
                    logging.info(f"Cloning all branches for {repo_name}...")
                    retry_subprocess(["git", "clone", "--mirror", azure_repo_url, repo_dir])
                else:
                    logging.info(f"Cloning default branch for {repo_name}...")
                    retry_subprocess(["git", "clone", azure_repo_url, repo_dir])
                logging.info(f"Cloning completed for TFS repo: {repo_name}")
            else:
                logging.info(f"Using existing local clone for {repo_name}")

            # If the repo contains large files and the flag is set, run BFG before push
            if has_large_files:
                logging.info(f"Repository {repo_name} marked for large file processing. Running BFG...")
                try:
                    process_large_files_with_bfg(repo_dir)
                    logging.info(f"BFG processing completed for {repo_name}")
                except Exception as bfg_err:
                    logging.error(f"BFG processing failed for {repo_name}: {bfg_err}")
                    # Continue or fail depending on desired behavior; here we fail to ensure visibility
                    raise

            if only_clone_repos:
                logging.info(f"Clone-only mode enabled. Skipping push for {repo_name}")
                return

            os.chdir(repo_dir)
            github_repo_url = f"https://{github_token}@github.com/{config['github_organization']}/{quote(git_repo)}.git"

            try:
                if specific_branches == "all":
                    logging.info(f"Pushing all branches and tags for {repo_name}...")
                    retry_subprocess(["git", "push", "--mirror", github_repo_url], enable_proxy=enable_proxy)
                else:
                    logging.info(f"Pushing selected branches: {specific_branches} for {repo_name}")
                    migrate_specific_branches(repo_dir, github_repo_url, specific_branches, enable_proxy=enable_proxy)
            except Exception as push_error:
                if "exceeds github’s file size limit" in str(push_error).lower() or "100mb" in str(push_error).lower():
                    raise Exception(f"Push failed due to large file(s): {push_error}")
                raise Exception(f"Push failed: {push_error}")

            save_checkpoint(collection_name, azure_project, repo_name)
            migration_summary.setdefault("migrated_repos", []).append(repo_name)
            logging.info(f"Migration successful for: {repo_name}")

            os.chdir(original_cwd)

            # Move BFG report if present at the requested source path
            move_bfg_report_to_log(azure_org, azure_project, repo_name)

            # Delete the local clone after successful migration
            try:
                shutil.rmtree(repo_dir)
                logging.info(f"Deleted cloned repo folder: {repo_dir}")
            except Exception as e:
                logging.warning(f"Failed to delete repo folder {repo_dir}: {e}")

    except Exception as e:
        logging.error(f"Error processing repo {tfs_repo}: {e}")
        migration_summary.setdefault("failed_repos", []).append({"repo": tfs_repo, "error": str(e)})
        raise
    finally:
        os.chdir(original_cwd)

def encrypt_secret(public_key: str, secret_value: str) -> str:
    """Encrypt a secret using the repository's public key (NaCl box encryption)."""
    public_key = nacl.public.PublicKey(public_key.encode("utf-8"), nacl.encoding.Base64Encoder())
    sealed_box = nacl.public.SealedBox(public_key)
    encrypted = sealed_box.encrypt(secret_value.encode("utf-8"))
    return base64.b64encode(encrypted).decode("utf-8")

def add_nexus_secrets(org, repo, github_token, nexus_username, nexus_password):
    try:
        headers = {
            "Authorization": f"token {github_token}",
            "Accept": "application/vnd.github+json"
        }

        # Step 1: Get the public key for the repo
        url_key = f"https://api.github.com/repos/{org}/{repo}/actions/secrets/public-key"
        response = requests.get(url_key, headers=headers)
        if response.status_code != 200:
            raise Exception(f"Failed to fetch public key: {response.status_code} - {response.text}")
        public_key_data = response.json()
        public_key = public_key_data["key"]
        key_id = public_key_data["key_id"]

        # Step 2: Encrypt and upload each secret
        for name, value in {
            "NEXUS_USERNAME": nexus_username,
            "NEXUS_PASSWORD": nexus_password
        }.items():
            encrypted_value = encrypt_secret(public_key, value)
            url_secret = f"https://api.github.com/repos/{org}/{repo}/actions/secrets/{name}"
            payload = {
                "encrypted_value": encrypted_value,
                "key_id": key_id
            }
            response = requests.put(url_secret, headers=headers, json=payload)
            if response.status_code not in [201, 204]:
                raise Exception(f"Failed to add secret {name}: {response.status_code} - {response.text}")
            logging.info(f"Successfully added secret '{name}' to {org}/{repo}")

    except Exception as e:
        logging.error(f"Error adding NEXUS secrets: {e}")
        raise

def get_repo_id(org, repo, github_token):
    headers = {"Authorization": f"token {github_token}", "Accept": "application/vnd.github+json"}
    url = f"https://api.github.com/repos/{org}/{repo}"
    response = requests.get(url, headers=headers)
    if response.status_code == 200:
        return response.json().get("id")
    else:
        raise Exception(f"Failed to fetch repo ID: {response.status_code} - {response.text}")

def link_ucd_token_secret(org, repo, github_token):
    try:
        headers = {"Authorization": f"token {github_token}", "Accept": "application/vnd.github+json"}
        repo_id = get_repo_id(org, repo, github_token)
        url = f"https://api.github.com/orgs/{org}/actions/secrets/UCD_ACCESS_TOKEN/repositories/{repo_id}"
        response = requests.put(url, headers=headers)
        if response.status_code not in [204, 201]:
            raise Exception(f"Failed to link UCD_ACCESS_TOKEN secret: {response.status_code} - {response.text}")
        logging.info(f"Linked UCD_ACCESS_TOKEN to {org}/{repo}")
    except Exception as e:
        logging.error(f"Error linking UCD_ACCESS_TOKEN: {e}")
        raise

def add_repo_to_runner_group(org, repo, github_token):
    try:
        headers = {"Authorization": f"token {github_token}", "Accept": "application/vnd.github+json"}
        response = requests.get(f"https://api.github.com/orgs/{org}/actions/runner-groups", headers=headers)
        runner_groups = response.json().get("runner_groups", [])
        group = next((g for g in runner_groups if g["name"] == "Migration-Runners"), None)
        if not group:
            raise Exception("Runner group 'Migration-Runners' not found")

        group_id = group["id"]
        repo_id = get_repo_id(org, repo, github_token)
        url = f"https://api.github.com/orgs/{org}/actions/runner-groups/{group_id}/repositories/{repo_id}"
        response = requests.put(url, headers=headers)
        if response.status_code not in [204, 201]:
            raise Exception(f"Failed to add repo to runner group: {response.status_code} - {response.text}")
        logging.info(f"Added {repo} to Migration-Runners runner group")
    except Exception as e:
        logging.error(f"Error adding repo to runner group: {e}")
        raise

def create_feature_pipeline_branch(org, repo, base_branch, github_token):
    try:
        headers = {"Authorization": f"token {github_token}", "Accept": "application/vnd.github+json"}
        url = f"https://api.github.com/repos/{org}/{repo}/git/ref/heads/{base_branch}"
        response = requests.get(url, headers=headers)
        if response.status_code != 200:
            raise Exception(f"Failed to get base branch: {response.status_code} - {response.text}")
        sha = response.json()["object"]["sha"]
        payload = {"ref": "refs/heads/feature/tfs-github-migration-pipeline", "sha": sha}
        response = requests.post(f"https://api.github.com/repos/{org}/{repo}/git/refs", headers=headers, json=payload)
        if response.status_code not in [201, 200]:
            raise Exception(f"Failed to create feature branch: {response.status_code} - {response.text}")
        logging.info(f"Created branch feature/tfs-github-migration-pipeline from {base_branch}")
    except Exception as e:
        logging.error(f"Error creating feature branch: {e}")
        raise

def update_excel_repo_id_column(excel_file, tfs_repo, repo_id):
    wb = load_workbook(excel_file)
    ws = wb["migration"]
    headers = [cell.value for cell in ws[1]]
    if "REPO_ID" not in headers:
        ws.cell(row=1, column=ws.max_column + 1, value="REPO_ID")
        repo_id_col = ws.max_column
    else:
        repo_id_col = headers.index("REPO_ID") + 1

    for row in ws.iter_rows(min_row=2):
        if row[headers.index("TFS_REPO")].value and str(row[headers.index("TFS_REPO")].value).strip().lower() == tfs_repo.strip().lower():
            row[repo_id_col - 1].value = repo_id
            break

    wb.save(excel_file)
    logging.info(f"Updated Excel with REPO_ID for {tfs_repo}")

def run_migration_for_all_configs(excel_file, template_file, output_folder, only_clone_repos, enable_proxy=False, token_file=None):
    config_files, df, row_mapping = generate_config_jsons(excel_file, template_file, output_folder)
    migration_results = {}
    comment_results = {}
    successful_repos = []

    # Load token data inside this function
    token_data = load_tokens(token_file)
    nexus_username = token_data.get("NEXUS_USERNAME")
    nexus_password = token_data.get("NEXUS_PASSWORD")

    for config_path in config_files:
        config = load_config_from_path(config_path)
        idx = row_mapping[config_path]
        repo_name = str(config.get("tfs_source_repo", "Unknown")).strip()
        org = config.get("github_organization", "")
        gh_repo = config.get("github_target_repo", "")
        github_token = config.get("github_token", "")
        tfs_token = config.get("azure_devops_pat_token", "")
        collection = config.get("azure_devops_organization", "")
        project = config.get("azure_devops_project", "")
        error_message = None

        # Setup per-repo log file
        log_base_dir = os.path.join(os.getcwd(), "tfs_repo_migration", "tfs_git_migration_log")
        collection_log_dir = os.path.join(log_base_dir, f"{collection}_{project}")
        os.makedirs(collection_log_dir, exist_ok=True)
        timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        log_path = os.path.join(collection_log_dir, f"{collection}_{project}_{timestamp}.log")

        log_handler = logging.FileHandler(log_path, mode='w', encoding='utf-8')
        log_handler.setFormatter(logging.Formatter("%(asctime)s [%(levelname)s] %(message)s"))
        logging.getLogger().addHandler(log_handler)

        logging.info(f"--- Started migration for {collection}/{project}/{repo_name} ---")

        try:
            # Step 1: Populate and set default branch
            try:
                populate_default_branch_from_tfs(
                    excel_file,
                    tfs_base_url=f"http://{config['tfs_url']}",
                    tfs_token=tfs_token
                )
            except Exception as e:
                error_message = f"Default branch details collection from TFS failed: {str(e)}"
                raise

            # Step 2: Clone from TFS
            try:
                clone_and_push_repositories(config, only_clone_repos=True, enable_proxy=enable_proxy)
            except Exception as e:
                error_message = f"Clone failed: {str(e)}"
                raise

            # Step 3: Create GitHub repo
            try:
                headers = {
                    "Authorization": f"token {github_token}",
                    "Accept": "application/vnd.github+json"
                }
                url = f"https://api.github.com/orgs/{org}/repos"
                payload = {"name": gh_repo, "auto_init": False, "private": True}
                response = github_request("POST", url, headers=headers, json=payload, enable_proxy=enable_proxy)

                if response.status_code == 201:
                    logging.info(f"Created GitHub repo: {gh_repo}")
                elif response.status_code == 422 and "already exists" in response.text:
                    logging.info(f"GitHub repo already exists: {gh_repo}")
                else:
                    error_message = f"GitHub repo creation failed: {response.status_code} - {response.text}"
                    raise Exception(error_message)
            except Exception as e:
                error_message = f"GitHub repo creation failed: {str(e)}"
                raise

            # Step 4: Push to GitHub
            try:
                clone_and_push_repositories(config, only_clone_repos=False, enable_proxy=enable_proxy)
            except Exception as e:
                error_message = f"Push failed: {str(e)}"
                raise

            # Step 5: Set default branch
            try:
                df_updated = pd.read_excel(excel_file, sheet_name="migration")
                updated_row = df_updated[df_updated["TFS_REPO"].str.strip().str.lower() == repo_name.lower()]
                branch = updated_row.iloc[0].get("DEFAULT_BRANCH", "").strip()

                if branch:
                    success = False
                    for i in range(3):
                        try:
                            set_default_branch(org, gh_repo, branch, github_token, enable_proxy=enable_proxy)
                            success = True
                            break
                        except Exception as e:
                            logging.warning(f"Retry {i+1}/3: Failed to set default branch for {gh_repo}: {e}")
                            time.sleep(2 ** i)
                    if not success:
                        error_message = f"Failed to set default branch after retries"
                        raise Exception(error_message)
                else:
                    error_message = f"DEFAULT_BRANCH missing for {gh_repo}"
                    raise Exception(error_message)
            except Exception as e:
                error_message = f"Set default branch failed: {str(e)}"
                raise

            # Step 6: Assign users/teams
            try:
                success = False
                for i in range(3):
                    try:
                        assign_repo_access_from_access_sheet(
                            excel_file, org, gh_repo, github_token, enable_proxy=enable_proxy
                        )
                        success = True
                        break
                    except Exception as e:
                        logging.warning(f"Retry {i+1}/3: Failed to assign access for {gh_repo}: {e}")
                        time.sleep(2 ** i)
                if not success:
                    error_message = f"Failed to assign access after retries"
                    raise Exception(error_message)
            except Exception as e:
                error_message = f"Assign access failed: {str(e)}"
                raise

            # Step 7: Post-access enhancements
            try:
                df_updated = pd.read_excel(excel_file, sheet_name="migration")
                updated_row = df_updated[df_updated["TFS_REPO"].str.strip().str.lower() == repo_name.lower()]

                nexus_val = updated_row.iloc[0].get("NEXUS", "").strip().lower()
                raw_feature_branch = updated_row.iloc[0].get("FEATURE_BRANCH", "")
                feature_branch = str(raw_feature_branch).strip() if pd.notna(raw_feature_branch) else ""

                # Check UCD_TOKEN column
                ucd_val = str(updated_row.iloc[0].get("UCD_TOKEN", "")).strip().lower()

                if nexus_val == "yes":
                    add_nexus_secrets(org, gh_repo, github_token, nexus_username, nexus_password)

                # Only link UCD_TOKEN if Excel says YES
                if ucd_val == "yes":
                    link_ucd_token_secret(org, gh_repo, github_token)
                    
                if feature_branch:
                    create_feature_pipeline_branch(org, gh_repo, feature_branch, github_token)
                    add_repo_to_runner_group(org, gh_repo, github_token)

                repo_id = get_repo_id(org, gh_repo, github_token)
                update_excel_repo_id_column(excel_file, repo_name, repo_id)

            except Exception as e:
                error_message = f"Post-access enhancement failed: {str(e)}"
                raise

            # Success
            migration_results[idx] = ("Migration successful", "green")
            comment_results[idx] = ""
            successful_repos.append(repo_name)

        except Exception as e:
            logging.error(f"Migration failed for {repo_name}: {e}")
            migration_results[idx] = ("Migration failed", "red")
            comment_results[idx] = error_message or str(e)

        finally:
            logging.info(f"--- Completed migration for {collection}/{project}/{repo_name} ---")
            logging.getLogger().removeHandler(log_handler)
            log_handler.close()

    update_excel_with_status(excel_file, migration_results, comment_results)
    return successful_repos

def assign_repo_access_from_access_sheet(excel_file, org_name, repo_name, github_token, enable_proxy=False):
    try:
        df = pd.read_excel(excel_file, sheet_name="access")
    except Exception as e:
        logging.warning(f"No access sheet or failed to read it: {e}")
        return

    df.columns = [str(col).strip().upper() for col in df.columns]
    g = Github(github_token)

    repo_rows = df[
        (df["GITHUB_ORGANIZATION"].str.strip().str.lower() == org_name.lower()) &
        (df["GITHUB_REPO"].str.strip().str.lower() == repo_name.lower())
    ]

    if repo_rows.empty:
        return

    try:
        org = g.get_organization(org_name)
        repo = org.get_repo(repo_name)
    except GithubException as e:
        logging.error(f"Failed to access repo '{repo_name}' in org '{org_name}': {e}")
        return

    for _, row in repo_rows.iterrows():
        role = str(row.get('ROLE', '')).strip().lower()
        usernames = str(row.get('GITHUB_USERNAME', '')).split(',') if pd.notna(row.get('GITHUB_USERNAME')) else []
        teams = str(row.get('GITHUB_TEAM', '')).split(',') if pd.notna(row.get('GITHUB_TEAM')) else []

        for username in [u.strip() for u in usernames if u.strip()]:
            try:
                repo.add_to_collaborators(username, permission=role)
                logging.info(f"Assigned user '{username}' to '{repo_name}' with '{role}'")
            except GithubException as e:
                logging.error(f"Failed to assign user '{username}' to '{repo_name}': {e}")

        for team_name in [t.strip() for t in teams if t.strip()]:
            team = validate_team(org, team_name)
            if team:
                try:
                    team.add_to_repos(repo)
                    team.update_team_repository(repo, permission=role)
                    logging.info(f"Assigned team '{team_name}' to '{repo_name}' with '{role}'")
                except GithubException as e:
                    logging.error(f"Failed to assign team '{team_name}' to '{repo_name}': {e}")

def update_excel_with_status(excel_file, migration_results, comment_results):
    wb = load_workbook(excel_file)
    ws = wb["migration"]

    # Find or create MIGRATION_STATUS column
    headers = [cell.value for cell in ws[1]]
    if "MIGRATION_STATUS" not in headers:
        ws.cell(row=1, column=ws.max_column + 1, value="MIGRATION_STATUS")
        status_col = ws.max_column
    else:
        status_col = headers.index("MIGRATION_STATUS") + 1

    # Find or create COMMENTS column
    headers = [cell.value for cell in ws[1]]
    if "COMMENTS" not in headers:
        ws.cell(row=1, column=ws.max_column + 1, value="COMMENTS")
        comment_col = ws.max_column
    else:
        comment_col = headers.index("COMMENTS") + 1

    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    for row_index, (status, color) in migration_results.items():
        ws.cell(row=row_index + 2, column=status_col, value=status).fill = green_fill if color == "green" else red_fill
        ws.cell(row=row_index + 2, column=comment_col, value=comment_results.get(row_index, ""))

    wb.save(excel_file)
    logging.info("Excel file updated with status and comments.")

def populate_default_branch_from_tfs(excel_file, tfs_base_url, tfs_token):
    # Load the Excel file and read the 'migration' sheet
    df = pd.read_excel(excel_file, sheet_name="migration")
 
    # Add 'DEFAULT_BRANCH' column if it doesn't exist
    if 'DEFAULT_BRANCH' not in df.columns:
        df['DEFAULT_BRANCH'] = ''
 
    for index, row in df.iterrows():
        collection = row.get('COLLECTION_NAME')
        project = row.get('PROJECT_NAME')
        repo = row.get('TFS_REPO')
 
        # Skip rows with missing or invalid data
        if not collection or not project or not repo:
            continue
 
        # Skip if 'DEFAULT_BRANCH' is already populated
        if pd.notna(row.get('DEFAULT_BRANCH')) and str(row.get('DEFAULT_BRANCH')).strip():
            continue
 
        # Encode inputs to ensure they are URL-safe
        encoded_collection = quote(str(collection))
        encoded_project = quote(str(project))
        encoded_repo = quote(str(repo))
 
        # Construct the API URL
        api_url = f"{tfs_base_url}/{encoded_collection}/{encoded_project}/_apis/git/repositories/{encoded_repo}?api-version=5.0"
 
        # Set up headers for the API request
        headers = {
            "Authorization": f"Basic {base64.b64encode(f':{tfs_token}'.encode()).decode()}",
            "Content-Type": "application/json"
        }
 
        try:
            # Make the API request
            response = requests.get(api_url, headers=headers)
            if response.status_code == 200:
                data = response.json()
                default_branch = data.get("defaultBranch", "")
 
                # Remove "refs/heads/" prefix if present
                if default_branch.startswith("refs/heads/"):
                    default_branch = default_branch.replace("refs/heads/", "")
 
                # Update the DataFrame with the default branch
                df.at[index, 'DEFAULT_BRANCH'] = default_branch
                logging.info(f"Set {default_branch} as default branch for {repo}")
            else:
                logging.warning(f"Failed to fetch default branch for {repo}: {response.status_code} - {response.text}")
        except Exception as e:
            logging.error(f"Exception fetching default branch for {repo}: {e}")
 
    # Update the Excel file with the modified DataFrame
    safe_update_sheet(excel_file, 'migration', df)
    logging.info("Excel file updated with default branches from TFS.")

def set_default_branch(org, repo, branch, github_token, enable_proxy=False):
    url = f"https://api.github.com/repos/{org}/{repo}"
    headers = {
        "Authorization": f"token {github_token}",
        "Accept": "application/vnd.github+json"
    }
    data = {"default_branch": branch}
    response = github_request("PATCH", url, headers=headers, json=data, enable_proxy=enable_proxy)

    if response:
        if response.status_code == 200:
            logging.info(f"Default branch for {org}/{repo} set to '{branch}'")
        else:
            logging.warning(f"Failed to update {org}/{repo}: {response.status_code} - {response.text}")

def process_default_branches(excel_file, github_token, enable_proxy=False):
    try:
        df = pd.read_excel(excel_file, sheet_name="migration")
    except Exception as e:
        logging.error(f"Error reading Excel file for default branches: {e}")
        return

    required_columns = {'GITHUB_REPO', 'GITHUB_ORGANIZATION', 'DEFAULT_BRANCH'}
    if not required_columns.issubset(df.columns):
        logging.warning(f"Excel must contain columns: {required_columns} to set default branches.")
        return

    for index, row in df.iterrows():
        org = row['GITHUB_ORGANIZATION']
        repo = row['GITHUB_REPO']
        branch = row['DEFAULT_BRANCH']
        set_default_branch(org, repo, branch, github_token, enable_proxy=enable_proxy)

def validate_team(org, team_name):
    for team in org.get_teams():
        if team.name.lower() == team_name.lower():
            return team
    return None

def validate_user(g, username):
    try:
        return g.get_user(username)
    except GithubException:
        return None

def assign_access(org, repo_name, team_name, username, role, g):
    try:
        repo = org.get_repo(repo_name)
    except GithubException:
        logging.error(f"Repository '{repo_name}' not found in organization '{org.login}'")
        return

    if pd.notna(team_name):
        team = validate_team(org, team_name)
        if team:
            try:
                team.add_to_repos(repo)
                team.update_team_repository(repo, permission=role.lower())
                logging.info(f"Added team '{team_name}' to '{repo_name}' with role '{role}'")
            except GithubException as e:
                logging.error(f"Failed to assign team '{team_name}' to '{repo_name}': {e}")
        else:
            logging.warning(f"Team '{team_name}' not found in organization '{org.login}'")

    elif pd.notna(username):
        user = validate_user(g, username)
        if user:
            try:
                repo.add_to_collaborators(username, permission=role.lower())
                logging.info(f"Added user '{username}' to '{repo_name}' with role '{role}'")
            except GithubException as e:
                logging.error(f"Failed to assign user '{username}' to '{repo_name}': {e}")
        else:
            logging.warning(f"User '{username}' not found on GitHub")
    else:
        logging.warning(f"No valid team or username provided for repo '{repo_name}'")

def process_access_assignment(excel_file, github_token, enable_proxy=False):
    try:
        df = pd.read_excel(excel_file, sheet_name="access")
    except Exception as e:
        logging.error(f"Failed to read Excel file '{excel_file}': {e}")
        return

    df.columns = [str(col).strip().upper() for col in df.columns]
    required_columns = {'GITHUB_REPO', 'GITHUB_ORGANIZATION', 'ROLE'}
    if not required_columns.issubset(df.columns):
        logging.error(f"Excel is missing one or more required columns: {required_columns}")
        return

    if 'GITHUB_USERNAME' not in df.columns and 'GITHUB_TEAM' not in df.columns:
        logging.warning("No GITHUB_USERNAME or GITHUB_TEAM column found; skipping access assignment.")
        return

    g = Github(github_token, per_page=100, retry=3, timeout=30)

    for idx, row in df.iterrows():
        repo_name = str(row.get('GITHUB_REPO', '')).strip()
        org_name = str(row.get('GITHUB_ORGANIZATION', '')).strip()
        role = str(row.get('ROLE', '')).strip().lower()
        usernames = str(row.get('GITHUB_USERNAME', '')).split(',') if pd.notna(row.get('GITHUB_USERNAME')) else []
        teams = str(row.get('GITHUB_TEAM', '')).split(',') if pd.notna(row.get('GITHUB_TEAM')) else []

        if not repo_name or not org_name or not role:
            logging.warning(f"Row {idx + 2} missing required fields. Skipping.")
            continue

        try:
            if org_name.lower() == g.get_user().login.lower():
                repo = g.get_user().get_repo(repo_name)
                org = None
            else:
                org = g.get_organization(org_name)
                repo = org.get_repo(repo_name)
        except GithubException as e:
            logging.error(f"Failed to access org '{org_name}' or repo '{repo_name}': {e}")
            continue

        for team_name in [t.strip() for t in teams if t.strip()]:
            try:
                team = next((t for t in org.get_teams() if t.name.lower() == team_name.lower()), None)
                if team:
                    team.add_to_repos(repo)
                    team.update_team_repository(repo, permission=role)
                    logging.info(f"Assigned team '{team_name}' to '{repo_name}' with '{role}'")
                else:
                    logging.warning(f"Team '{team_name}' not found in org '{org_name}'")
            except GithubException as e:
                logging.error(f"Error assigning team '{team_name}' to '{repo_name}': {e}")

        for username in [u.strip() for u in usernames if u.strip()]:
            try:
                repo.add_to_collaborators(username, permission=role)
                logging.info(f"Assigned user '{username}' to '{repo_name}' with '{role}'")
            except GithubException as e:
                logging.error(f"Error assigning user '{username}' to '{repo_name}': {e}")

        safe_update_sheet(excel_file, 'access', df)

# GitHub Repo Creation First
def create_repos_from_excel(excel_file, github_token, enable_proxy=False):
    headers = {
        "Authorization": f"token {github_token}",
        "Accept": "application/vnd.github+json"
    }
    df = pd.read_excel(excel_file, sheet_name="migration", usecols=["GITHUB_ORGANIZATION", "GITHUB_REPO"])
    df = df.dropna()
    print('started creating repos....')
    for _, row in df.iterrows():
        ignore_val = str(row.get("IGNORE", "")).strip().lower()
        if ignore_val == "yes":
            continue
        org = row["GITHUB_ORGANIZATION"]
        repo_name = row["GITHUB_REPO"]
        url = f"https://api.github.com/orgs/{org}/repos"
        payload = {"name": repo_name, "auto_init": False, "private": True}
        response = github_request("POST", url, headers=headers, json=payload, enable_proxy=enable_proxy)
        print(response.status_code, response.text)
        if response:
            if response.status_code == 201:
                print(f"Repository '{repo_name}' created under organization '{org}'")
            elif response.status_code == 422 and "already exists" in response.text:
                print(f"Repository '{repo_name}' already exists in '{org}'")
            else:
                print(f"Failed to create repo '{repo_name}' under '{org}': {response.status_code} - {response.text}")

def safe_update_sheet(excel_file, sheet_name, updated_df):
    # Load the workbook
    wb = load_workbook(excel_file)
 
    # Remove the existing sheet with that name (if it exists)
    if sheet_name in wb.sheetnames:
        std = wb[sheet_name]
        wb.remove(std)
 
    # Create a new sheet with the same name
    ws = wb.create_sheet(sheet_name)
 
    # Write the DataFrame to the new sheet
    for r_idx, row in enumerate(dataframe_to_rows(updated_df, index=False, header=True), start=1):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=value)
 
    # Save the workbook
    wb.save(excel_file)

def safe_update_sheet(file_path, sheet_name, df):
    # Load the existing workbook
    wb = load_workbook(file_path)
 
    # Remove the existing sheet if it exists
    if sheet_name in wb.sheetnames:
        std = wb[sheet_name]
        wb.remove(std)
 
    # Create a new sheet with the same name
    ws = wb.create_sheet(title=sheet_name)
 
    # Write the updated DataFrame to the sheet
    for row in dataframe_to_rows(df, index=False, header=True):
        ws.append(row)
 
    # Save the workbook
    wb.save(file_path)

def github_request(method, url, headers, json=None, enable_proxy=False):
    try:
        response = requests.request(
            method, url, headers=headers, json=json,
            proxies=PROXIES if enable_proxy else None
        )
        return response
    except requests.exceptions.RequestException as e:
        logging.error(f"GitHub request failed: {e}")
        return None

def get_tfs_domain_from_template(template_file):
    if not os.path.exists(template_file):
        raise FileNotFoundError(f"Template config file not found: {template_file}")

    with open(template_file, "r") as f:
        config = json.load(f)

    tfs_url = config.get("tfs_url", "").strip()
    return tfs_url

def run_command(command, cwd=None):
    try:
        result = subprocess.run(
            command,
            cwd=cwd,
            check=True,
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            text=True
        )
        logging.info(f"Command output:\n{result.stdout}")
        if result.stderr:
            logging.warning(f"Command stderr:\n{result.stderr}")
        return result.stdout
    except subprocess.CalledProcessError as e:
        logging.error(f"Command failed: {e.stderr.strip()}")
        raise

# --- Function to Process Large Files Using BFG ---
def process_large_files_with_bfg(repo_dir):
    logging.info(f"Processing large files in repository: {repo_dir}")

    # Step 1: Run BFG to remove large files
    bfg_command = ["java", "-jar", BFG_JAR_PATH, "--strip-blobs-bigger-than", "100M", "--no-blob-protection", repo_dir]
    run_command(bfg_command)

    # Step 2: Clean and repack the repository
    logging.info("Cleaning and repacking the repository...")
    run_command(["git", "reflog", "expire", "--expire=now", "--all"], cwd=repo_dir)
    run_command(["git", "gc", "--prune=now", "--aggressive"], cwd=repo_dir)

    logging.info(f"Large file processing completed for repository: {repo_dir}")

def main():
    parser = argparse.ArgumentParser(description="TFS (GIT) to GitHub migration script.")
    parser.add_argument("--only_clone_repos", action="store_true", help="Only clone repositories without pushing to GitHub.")
    parser.add_argument("--enable_proxy", action="store_true", help="Enable proxy for GitHub operations")
    parser.add_argument("--token_file", type=str, help="Path to JSON file containing both GitHub and TFS tokens.")
    parser.add_argument("--excel_file", type=str, required=True, help="Path to the Excel file for migration and access.")
    args = parser.parse_args()

    # Define paths
    excel_file = args.excel_file
    template_file = 'config.json'
    output_folder = 'generated_config_json'

    global github_pat, tfs_pat, tfs_domain
    tfs_domain = get_tfs_domain_from_template(template_file)

    # Load and validate tokens
    token_data = load_tokens(token_file=args.token_file)
    github_pat = token_data["github"]
    tfs_pat = token_data["tfs"]

    # Run migration (includes all steps per-repo now)
    try:
        run_migration_for_all_configs(
            excel_file, template_file, output_folder,
            args.only_clone_repos, enable_proxy=args.enable_proxy, token_file=args.token_file
        )
        save_summary_log()
    except Exception as e:
        logging.error(f"Migration failed: {e}")
        save_summary_log()

if __name__ == "__main__":
    main()
