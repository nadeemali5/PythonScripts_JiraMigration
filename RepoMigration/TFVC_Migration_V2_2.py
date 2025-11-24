import os
import subprocess
import requests
import json
import pandas as pd
import logging
from datetime import datetime
import base64
import time
from urllib.parse import quote
import shutil
import openpyxl
from openpyxl.styles import PatternFill
import nacl.encoding
import nacl.public
import argparse 
from github import Github, Auth

CHECKPOINT_FILE = os.path.join(os.getcwd(), "tfs_tfvc_checkpoint.json")

proxies = {
    "http": "http://webproxy.intranet.umb.com:3128",
    "https": "http://webproxy.intranet.umb.com:3128"
}

def setup_logging():
    """Sets up logging to console and a timestamped file."""
    try:
        base_dir = os.getcwd()
        log_dir = os.path.join(base_dir, "log")
        
        # Attempt to create the log directory
        try:
            os.makedirs(log_dir, exist_ok=True)
        except OSError as e:
            print(f"Error creating log directory: {e}")
            raise

        log_filename = datetime.now().strftime("TFVC_Migration_V1_1_log_%Y-%m-%d_%H-%M-%S.log")
        log_path = os.path.join(log_dir, log_filename)

        # Remove existing handlers if any
        for handler in logging.root.handlers[:]:
            logging.root.removeHandler(handler)

        logging.basicConfig(
            level=logging.INFO,
            format="%(asctime)s [%(levelname)s]: %(message)s",
            handlers=[
                logging.FileHandler(log_path, mode='w'),
                logging.StreamHandler()
            ]
        )
        logging.info(f"==================== TFVC Migration Started ====================")
        logging.info(f"Log file created at: {log_path}")
        return log_path

    except Exception as e:
        print(f"An error occurred while setting up logging: {e}")
        raise

def update_excel_with_status(excel_file, migration_results):
    """ Updates the input Excel file with migration status and comments for each repository."""
    try:
        # Load the Excel workbook
        try:
            wb = openpyxl.load_workbook(excel_file)
            sheet = wb.active
        except Exception as e:
            logging.error(f"Error loading Excel file: {e}")
            return

        # Add new columns for "MIGRATION STATUS" and "COMMENTS" if they don't already exist
        headers = [cell.value for cell in sheet[1]]

        # Columns to add for workflow steps
        step_columns = ["JSON_CREATED", "REPO_CREATED", "MIGRATION_DONE", "ACCESS_ASSIGNED", "SECRETS_ADDED"]
        for col in step_columns:
            if col not in headers:
                sheet.cell(row=1, column=len(headers) + 1, value=col)
                headers.append(col)

        # Ensure MIGRATION STATUS and COMMENTS exist
        if "MIGRATION STATUS" not in headers:
            sheet.cell(row=1, column=len(headers) + 1, value="MIGRATION STATUS")
            headers.append("MIGRATION STATUS")
        if "COMMENTS" not in headers:
            sheet.cell(row=1, column=len(headers) + 1, value="COMMENTS")
            headers.append("COMMENTS")

        # Get column indices
        col_indices = {h: headers.index(h) + 1 for h in headers}

        # Define color fills for success and failure
        success_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Green
        failure_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Red

        # Iterate through the migration results and update the Excel sheet
        for result in migration_results:
            for row in range(2, sheet.max_row + 1):
                if (sheet.cell(row=row, column=headers.index("COLLECTION_NAME") + 1).value == result["COLLECTION_NAME"] and
                    sheet.cell(row=row, column=headers.index("PROJECT_NAME") + 1).value == result["PROJECT_NAME"] and
                    sheet.cell(row=row, column=headers.index("BRANCH_LIST") + 1).value == result["BRANCH_LIST"]):
                    # Update step checklist
                    for step in step_columns:
                        sheet.cell(row=row, column=col_indices[step],
                                   value="✅" if result.get(step, False) else "❌")
                        sheet.cell(row=row, column=col_indices[step]).fill = success_fill if result.get(step, False) else failure_fill

                    # Update overall migration status
                    status = result["STATUS"]
                    sheet.cell(row=row, column=col_indices["MIGRATION STATUS"], value=status)
                    sheet.cell(row=row, column=col_indices["MIGRATION STATUS"]).fill = success_fill if status == "SUCCESSFUL" else failure_fill

                    # Update comments
                    sheet.cell(row=row, column=col_indices["COMMENTS"], value=result.get("COMMENTS", ""))
                    break

        # Save the updated Excel file
        wb.save(excel_file)
        logging.info(f"Excel updated successfully with workflow checklist: {excel_file}")

    except Exception as e:
        logging.error(f"Error updating Excel file: {e}")

def prompt_git_tfs_log_path():
    """Prompts the user for the git-tfs log path or user ID and sets up the directory."""
    user_input = input("Enter the full path for git-tfs logs or your user ID: ").strip()
    if os.path.exists(user_input):
        git_tfs_log_path = user_input
    else:
        git_tfs_log_path = f"C:\\Users\\{user_input}\\AppData\\Local\\git-tfs"
        print(f"Generated path: {git_tfs_log_path}")
        print("Make sure this path exists.")
    
    if not os.path.exists(git_tfs_log_path):
        raise FileNotFoundError(f"The path {git_tfs_log_path} does not exist. Please verify and try again.")
    
    # Manage existing log files
    old_logs_path = os.path.join(git_tfs_log_path, "old")
    os.makedirs(old_logs_path, exist_ok=True)
    files_moved = False
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    for ext in [".log", ".txt"]:
        files_to_move = [file for file in os.listdir(git_tfs_log_path) if file.endswith(ext)]
        if files_to_move:
            for file in files_to_move:
                src_path = os.path.join(git_tfs_log_path, file)
                file_name, file_ext = os.path.splitext(file)
                new_file_name = f"{file_name}_{timestamp}{file_ext}"  # append timestamp
                dest_path = os.path.join(old_logs_path, new_file_name)

                shutil.move(src_path, dest_path)
                logging.info(f"Moved old {ext} file '{file}' to '{dest_path}'.")
            files_moved = True
            break  # Stop after finding and moving the first type of file
    
    if not files_moved:
        logging.info("No .log or .txt files found to move.")

    # Create TFVC_log folder and today's date folder
    tfvc_log_path = os.path.join(git_tfs_log_path, "TFVC_log")
    os.makedirs(tfvc_log_path, exist_ok=True)
    today_date_folder = datetime.now().strftime("%d-%m-%Y")
    date_folder_path = os.path.join(tfvc_log_path, today_date_folder)
    os.makedirs(date_folder_path, exist_ok=True)
    logging.info(f"Created log folder for today's date: {date_folder_path}")

    return git_tfs_log_path, date_folder_path

def load_tokens():
    """Loads tokens from token.json."""
    token_file = os.path.join(os.getcwd(), "token.json")
    try:
        with open(token_file, "r") as file:
            logging.info(f"Loading tokens from {token_file}")
            return json.load(file)
    except Exception as e:
        logging.error(f"Error reading token file: {e}")
        raise

def load_config(config_path=None):
    """Loads configuration from config.json."""
    if not config_path:
        config_path = os.path.join(os.getcwd(), "config.json")
    try:
        with open(config_path, "r") as config_file:
            logging.info(f"Loading configuration from {config_path}")
            return json.load(config_file)
    except Exception as e:
        logging.error(f"Error reading configuration file {config_path}: {e}")
        raise

def save_checkpoint(collection_name, project_name, branch_name):
    """Saves migrated branch name under a JSON hierarchy in tfs_tfvc_checkpoint.json."""
    checkpoint = {}
    if os.path.exists(CHECKPOINT_FILE):
        with open(CHECKPOINT_FILE, "r") as file:
            checkpoint = json.load(file)

    if collection_name not in checkpoint:
        checkpoint[collection_name] = {}

    if project_name not in checkpoint[collection_name]:
        checkpoint[collection_name][project_name] = []

    if branch_name not in checkpoint[collection_name][project_name]:
        checkpoint[collection_name][project_name].append(branch_name)
        with open(CHECKPOINT_FILE, "w") as file:
            json.dump(checkpoint, file, indent=4)
        logging.info(f"Checkpoint updated: {branch_name} saved under {collection_name}/{project_name}.")

def load_checkpoint():
    """Loads the tfs_tfvc_checkpoint.json file, creating it if it doesn't exist."""
    if not os.path.exists(CHECKPOINT_FILE):
        with open(CHECKPOINT_FILE, "w") as file:
            json.dump({}, file)  # Create an empty JSON object
        logging.info(f"Created new checkpoint file: {CHECKPOINT_FILE}")
    with open(CHECKPOINT_FILE, "r") as file:
        logging.info(f"Loading checkpoint from {CHECKPOINT_FILE}")
        return json.load(file)

def retry_subprocess(command, enable_proxy=False, retries=3, fatal_substrings=None,
    retryable_substrings=None,
    backoff_base=2,
    timeout=None, **kwargs):
    """Retries a subprocess command in case of network issues."""
    env = os.environ.copy()
    if enable_proxy:
        http_proxy = "http://webproxy.intranet.umb.com:3128"
        https_proxy = "http://webproxy.intranet.umb.com:3128"
        env['http_proxy'] = http_proxy
        env['https_proxy'] = https_proxy

    # Capture output for analysis unless caller provided their own pipes
    capture = "stdout" not in kwargs and "stderr" not in kwargs
    if capture:
        kwargs.setdefault("stdout", subprocess.PIPE)
        kwargs.setdefault("stderr", subprocess.PIPE)
        kwargs.setdefault("text", True)

    # Backward-compat: allow callers to pass check=False to ignore non-zero exit codes
    ignore_nonzero = False
    if "check" in kwargs:
        ignore_nonzero = kwargs.pop("check") is False

    # Default fatal indicators (do not retry)
    default_fatal = [
        "doesn't exist",
        "does not exist",
        "not found",
        "tf400",                # e.g., TF400324, TF400367
        "unable to access",
        "invalid path",
        "authentication failed",
    ]
    fatal_tokens = [s.lower() for s in (fatal_substrings or default_fatal)]

    # Default retryable indicators (transient issues)
    default_retryable = [
        "timed out",
        "timeout",
        "connection reset",
        "connection refused",
        "could not resolve host",
        "remote end hung up unexpectedly",
        "internal server error",
        "502",
        "503",
        "504",
        "rate limit",
        "tls handshake timeout",
        "proxy",
    ]
    retry_tokens = [s.lower() for s in (retryable_substrings or default_retryable)]

    for attempt in range(1, retries + 1):
        try:
            proc = subprocess.run(
                command,
                env=env,
                timeout=timeout,
                check=False,  # we evaluate rc ourselves
                **kwargs
            )
            rc = proc.returncode
            out = (proc.stdout or "").strip() if capture else ""
            err = (proc.stderr or "").strip() if capture else ""
            combined = (out + "\n" + err).lower()

            if rc == 0:
                logging.info("Command executed successfully")
                if out:
                    logging.debug(out)
                return True

            if ignore_nonzero:
                logging.info("Command returned non-zero exit code, but check=False was requested. Continuing: %s", " ".join(command))
                if out:
                    logging.debug(out)
                if err:
                    logging.debug(err)
                return True

            if capture and any(token in combined for token in fatal_tokens):
                logging.error("Fatal error detected; not retrying: %s", " ".join(command))
                if out:
                    logging.error(out)
                if err:
                    logging.error(err)
                return False

            should_retry = capture and any(token in combined for token in retry_tokens)

            if attempt < retries and (should_retry or not capture):
                delay = backoff_base ** (attempt - 1)
                logging.warning("Attempt %d/%d failed (rc=%d). Retrying in %ds: %s", attempt, retries, rc, delay, " ".join(command))
                if err:
                    logging.warning(err)
                time.sleep(delay)
                continue

            logging.error("Command failed (attempt %d/%d, rc=%d): %s", attempt, retries, rc, " ".join(command))
            if out:
                logging.error(out)
            if err:
                logging.error(err)
            return False

        except Exception as e:
            if attempt < retries:
                delay = backoff_base ** (attempt - 1)
                logging.warning("Exception running command (attempt %d/%d): %s. Retrying in %ds. Error: %s", attempt, retries, " ".join(command), delay, e)
                time.sleep(delay)
                continue
            logging.error("Command crashed after %d attempts: %s. Error: %s", retries, " ".join(command), e)
            return False

def get_github_auth_header(github_token):
    """Returns the authorization header for GitHub API requests."""
    return {"Authorization": f"token {github_token}"}

def check_github_repo_exists(github_org, repo_name, github_token):
    """Checks if a repository exists on GitHub."""
    url = f"https://api.github.com/repos/{github_org}/{repo_name}"
    headers = get_github_auth_header(github_token)
    logging.info(f"Checking if GitHub repository exists: {github_org}/{repo_name}")
    response = requests.get(url, headers=headers, proxies=proxies)
    logging.info(f"GitHub API response status: {response.status_code}")
    return response.status_code == 200

def github_request(method, url, headers, json=None, enable_proxy=False):
    """Handles GitHub API requests."""
    try:
        response = requests.request(
            method, url, headers=headers, json=json,
            proxies=proxies if enable_proxy else None
        )
        return response
    except requests.exceptions.RequestException as e:
        logging.error(f"GitHub request failed: {e}")
        return None

def create_repos_from_excel(row: dict, github_token: str, enable_proxy=False):
    """Creates repositories in GitHub based on the Excel file."""
    try:
        ignore_val = str(row.get("IGNORE", "") or "").strip().lower()
        if ignore_val == "yes":
            logging.info(f"Skipping repo creation for row: {row}")
            return

        org = str(row.get("GITHUB_ORGANIZATION", "") or "").strip()
        repo_name = str(row.get("GITHUB_REPO", "") or "").strip()

        if not org or not repo_name:
            logging.warning(f"Missing organization or repo name in row: {row}")
            return
        
        headers = {
            "Authorization": f"token {github_token}",
            "Accept": "application/vnd.github+json"
        }
        
        url = f"https://api.github.com/orgs/{org}/repos"
        payload = {"name": repo_name, "auto_init": False, "private": True}
        response = github_request("POST", url, headers=headers, json=payload, enable_proxy=enable_proxy)
        if response:
            if response.status_code == 201:
                logging.info(f"Repository '{repo_name}' created under organization '{org}'")
            elif response.status_code == 422 and "already exists" in response.text:
                logging.info(f"Repository '{repo_name}' already exists in '{org}'")
            else:
                logging.error(f"Failed to create repo '{repo_name}' under '{org}': {response.status_code} - {response.text}")
    except Exception as e:
        logging.error(f"Error creating repo from row: {e}")

def encrypt_secret(public_key: str, secret_value: str) -> str:
    """Encrypt a secret using the repository's public key (NaCl box encryption)."""
    public_key = nacl.public.PublicKey(public_key.encode("utf-8"), nacl.encoding.Base64Encoder())
    sealed_box = nacl.public.SealedBox(public_key)
    encrypted = sealed_box.encrypt(secret_value.encode("utf-8"))
    return base64.b64encode(encrypted).decode("utf-8")

def add_nexus_secrets(org, repo, github_token, nexus_username, nexus_password):
    """Adds Nexus secrets to the GitHub repository."""
    try:
        headers = {
            "Authorization": f"token {github_token}",
            "Accept": "application/vnd.github+json"
        }

        # Step 1: Get the public key for the repo
        url_key = f"https://api.github.com/repos/{org}/{repo}/actions/secrets/public-key"
        response = requests.get(url_key, headers=headers)
        if response.status_code != 200:
            raise Exception(f"Failed to fetch public key: {response.status_code} - {response.text}")
        public_key_data = response.json()
        public_key = public_key_data["key"]
        key_id = public_key_data["key_id"]

        # Step 2: Encrypt and upload each secret
        for name, value in {
            "NEXUS_USERNAME": nexus_username,
            "NEXUS_PASSWORD": nexus_password
        }.items():
            encrypted_value = encrypt_secret(public_key, value)
            url_secret = f"https://api.github.com/repos/{org}/{repo}/actions/secrets/{name}"
            payload = {
                "encrypted_value": encrypted_value,
                "key_id": key_id
            }
            response = requests.put(url_secret, headers=headers, json=payload)
            if response.status_code not in [201, 204]:
                raise Exception(f"Failed to add secret {name}: {response.status_code} - {response.text}")
            logging.info(f"Successfully added secret '{name}' to {org}/{repo}")
    except Exception as e:
        logging.error(f"Error adding NEXUS secrets: {e}")
        raise

def get_repo_id(org, repo, github_token):
    headers = {"Authorization": f"token {github_token}", "Accept": "application/vnd.github+json"}
    url = f"https://api.github.com/repos/{org}/{repo}"
    response = requests.get(url, headers=headers)
    if response.status_code == 200:
        return response.json().get("id")
    else:
        raise Exception(f"Failed to fetch repo ID: {response.status_code} - {response.text}")

def link_ucd_token_secret(org, repo, github_token):
    """Links the UCD_ACCESS_TOKEN secret to the GitHub repository."""
    try:
        headers = {"Authorization": f"token {github_token}", "Accept": "application/vnd.github+json"}
        repo_id = get_repo_id(org, repo, github_token)
        url = f"https://api.github.com/orgs/{org}/actions/secrets/UCD_ACCESS_TOKEN/repositories/{repo_id}"
        response = requests.put(url, headers=headers)
        if response.status_code not in [204, 201]:
            raise Exception(f"Failed to link UCD_ACCESS_TOKEN secret: {response.status_code} - {response.text}")
        logging.info(f"Linked UCD_ACCESS_TOKEN to {org}/{repo}")
    except Exception as e:
        logging.error(f"Error linking UCD_ACCESS_TOKEN: {e}")
        raise

def create_feature_pipeline_branch(org, repo, base_branch, github_token):
    """Creates a feature branch in the GitHub repository."""
    try:
        headers = {"Authorization": f"token {github_token}", "Accept": "application/vnd.github+json"}
        url = f"https://api.github.com/repos/{org}/{repo}/git/ref/heads/{base_branch}"
        response = requests.get(url, headers=headers)

        if response.status_code == 404:
            logging.warning(f"Base branch '{base_branch}' not found in repo {repo}. Skipping feature branch creation.")
            return False

        if response.status_code == 409:
            logging.warning(f"Repo '{repo}' is empty. Skipping feature branch creation until TFVC code is pushed.")
            return False

        if response.status_code != 200:
            raise Exception(f"Failed to get base branch: {response.status_code} - {response.text}")
        sha = response.json()["object"]["sha"]
        payload = {"ref": "refs/heads/feature/tfs-github-migration-pipeline", "sha": sha}
        response = requests.post(f"https://api.github.com/repos/{org}/{repo}/git/refs", headers=headers, json=payload)
        if response.status_code not in [201, 200]:
            raise Exception(f"Failed to create feature branch: {response.status_code} - {response.text}")
        logging.info(f"Created branch feature/tfs-github-migration-pipeline from {base_branch}")
        return True
    except Exception as e:
        logging.error(f"Error creating feature branch in repo '{repo}': {e}")
        return False

def process_excel_file_for_secrets(row, github_token, nexus_username, nexus_password):
    """Processes a single row from the Excel file to handle UCD_TOKEN, NEXUS, and FEATURE_BRANCH."""
    
    org = row.get("GITHUB_ORGANIZATION")
    repo = row.get("GITHUB_REPO")
    ucd_token = str(row.get("UCD_TOKEN", "") or "").strip().lower()
    nexus = str(row.get("NEXUS", "") or "").strip().lower()
    feature_branch = str(row.get("FEATURE_BRANCH", "") or "").strip()

    status_flags = {
        "UCD_TOKEN": False,
        "NEXUS": False,
        "FEATURE_BRANCH": False
    }
    try:
        if nexus == "yes":
            add_nexus_secrets(org, repo, github_token, nexus_username, nexus_password)
            status_flags["NEXUS"] = True
            logging.info(f"Nexus secrets added for {org}/{repo}")
        else:
            logging.info("No Nexus secrets to add")
        if ucd_token == "yes":
            link_ucd_token_secret(org, repo, github_token)
            status_flags["UCD_TOKEN"] = True
            logging.info(f"UCD_TOKEN secret linked for {org}/{repo}")
        else:
            logging.info("No ucd token to add")
        if feature_branch:
            create_feature_pipeline_branch(org, repo, feature_branch, github_token)
            add_repo_to_runner_group(org, repo, github_token)
            status_flags["FEATURE_BRANCH"] = True
            logging.info(f"Feature branch '{feature_branch}' pipeline configured for {org}/{repo}")
        else:
            logging.info("No feature branch to add")
    except Exception as e:
        logging.error(f"Error processing secrets for {org}/{repo}: {e}")

def validate_team(org, team_name):
    for team in org.get_teams():
        if team.name.lower() == team_name.lower():
            return team
    return None

def assign_repo_access_from_access_sheet(excel_file, org_name, repo_name, github_token):
    try:
        df = pd.read_excel(excel_file, sheet_name="access")
    except Exception as e:
        logging.warning(f"No access sheet or failed to read it: {e}")
        return

    df.columns = [str(col).strip().upper() for col in df.columns]
    g = Github(auth=Auth.Token(github_token))

    repo_rows = df[
        (df["GITHUB_ORGANIZATION"].str.strip().str.lower() == org_name.lower()) &
        (df["GITHUB_REPO"].str.strip().str.lower() == repo_name.lower())
    ]

    if repo_rows.empty:
        return

    try:
        org = g.get_organization(org_name)
        repo = org.get_repo(repo_name)
    except Exception as e:
        logging.error(f"Failed to access repo '{repo_name}' in org '{org_name}': {e}")
        return

    for _, row in repo_rows.iterrows():
        role = str(row.get('ROLE', '')).strip().lower()
        usernames = str(row.get('GITHUB_USERNAME', '')).split(',') if pd.notna(row.get('GITHUB_USERNAME')) else []
        teams = str(row.get('GITHUB_TEAM', '')).split(',') if pd.notna(row.get('GITHUB_TEAM')) else []

        for username in [u.strip() for u in usernames if u.strip()]:
            try:
                repo.add_to_collaborators(username, permission=role)
                logging.info(f"Assigned user '{username}' to '{repo_name}' with '{role}'")
            except Exception as e:
                logging.error(f"Failed to assign user '{username}' to '{repo_name}': {e}")

        for team_name in [t.strip() for t in teams if t.strip()]:
            team = validate_team(org, team_name)
            if team:
                try:
                    team.add_to_repos(repo)
                    team.update_team_repository(repo, permission=role)
                    logging.info(f"Assigned team '{team_name}' to '{repo_name}' with '{role}'")
                except Exception as e:
                    logging.error(f"Failed to assign team '{team_name}' to '{repo_name}': {e}")

def update_excel_repo_id_column(excel_file, branch_list_str, repo_id):
    """Updates the Excel file with the GitHub repo ID for the given TFS_REPO. Creates a REPO_ID column if it doesn't already exist."""
    try:
        wb = openpyxl.load_workbook(excel_file)
        sheet = wb.active

        # Get headers from first row
        headers = [cell.value for cell in sheet[1]]

        # Ensure REPO_ID column exists
        if "REPO_ID" not in headers:
            sheet.cell(row=1, column=len(headers) + 1, value="REPO_ID")
            headers.append("REPO_ID")

        # Get column indices
        col_indices = {h: headers.index(h) + 1 for h in headers}

        # Locate the correct row and update REPO_ID
        for row in range(2, sheet.max_row + 1):
            branch_list_value = sheet.cell(row=row, column=headers.index("BRANCH_LIST") + 1).value
            if branch_list_value and str(branch_list_value).strip().lower() == branch_list_str.strip().lower():
                sheet.cell(row=row, column=col_indices["REPO_ID"], value=repo_id)
                logging.info(f"Updated REPO_ID for BRANCH_LIST '{branch_list_str}' with {repo_id}")
                break
        # Save the workbook
        wb.save(excel_file)

    except Exception as e:
        logging.error(f"Error updating Excel with repo ID for BRANCH_LIST '{branch_list_str}': {e}")

def clone_and_push_tfvc(config, git_tfs_log_path, date_folder_path):
    """Clones TFVC repositories from TFS and pushes them to GitHub."""
    tfs_url = config["tfs_url"]
    collection_name = config["azure_devops_organization"]
    project_name = config["azure_devops_project"]
    github_org = config["github_organization"]
    github_repo_name = config["github_repo_name"]
    github_token = config["github_token"]
    specific_branches = config.get("specific_branches", [])
    github_master_branch = config["tfvc_master_branch"]

    # Create the main directory structure
    repo_base_dir = os.path.join(os.getcwd(), "s", collection_name, project_name)
    os.makedirs(repo_base_dir, exist_ok=True)
    logging.info(f"Repo base directory: {repo_base_dir}")

    logging.info("Starting TFVC to GitHub migration process...")

    # Check if the GitHub repository exists
    if not check_github_repo_exists(github_org, github_repo_name, github_token):
        logging.error(f"GitHub repository '{github_repo_name}' does not exist. Exiting...")
        return 0
    
    checkpoint = load_checkpoint()
    root_dir = os.getcwd()

    for branch_path in specific_branches:
        branch_name = branch_path.split("/")[-1]  # Extract branch name from TFVC path
        

        logging.info(f"Processing branch: {branch_name}")

        if collection_name in checkpoint and project_name in checkpoint[collection_name] and branch_name in checkpoint[collection_name][project_name]:
            logging.info(f"Skipping migration for branch {branch_name}, already migrated.")
            continue
            
            # Checklist for this branch
        success_flags = {
            "dir_created": False,
            "cloned": False,
            "labels_fetched": False,
            "remote_set": False,
            "pushed_branch": False,
            "pushed_tags": False,
            "logs_moved": False
        }
        try:
            # 1. Create branch directory
            branch_dir = os.path.join(repo_base_dir, branch_name)
            os.makedirs(branch_dir, exist_ok=True)
            success_flags["dir_created"] = True
            logging.info(f"Created directory for branch '{branch_name}': {branch_dir}")

            # Clone the TFVC repository into the branch directory
            logging.info(f"Cloning TFVC path '{branch_path}'...")
            if not retry_subprocess([
                "git-tfs", "clone",
                f"{tfs_url}/{collection_name}",
                branch_path,
                branch_dir,
                "--export", "--branches=none"
            ],
                fatal_substrings=[
                    "doesn't exist",
                    "does not exist",
                    "not found",
                    "tf400",
                    "unable to access",
                    "invalid path",
                    "authentication failed",
                ],):
                return 0

            # Change directory to the cloned repository
            os.chdir(branch_dir)
            logging.info(f"Changed directory to {branch_dir}")

             # Post-clone sanity check: ensure at least one commit exists
            if not retry_subprocess(["git", "rev-parse", "--verify", "HEAD"], retries=1):
                logging.error(f"Clone produced an empty or invalid repository for '{branch_name}'. Aborting.")
                os.chdir(root_dir)
                shutil.rmtree(branch_dir, ignore_errors=True)
                return 0

            success_flags["cloned"] = True

            # Fetch labels (tags) from TFVC
            logging.info(f"Fetching labels for branch '{branch_name}'...")
            if not retry_subprocess(["git-tfs", "labels", "--all"]):
                return 0
            success_flags["labels_fetched"] = True
            if not retry_subprocess(["git", "tag"]):
                return 0

            # 4. Set GitHub remote
            retry_subprocess(["git", "remote", "remove", "origin"], check=False)
            if not retry_subprocess(["git", "remote", "add", "origin", f"https://{github_token}@github.com/{github_org}/{github_repo_name}.git"], enable_proxy=True):
                return 0
            success_flags["remote_set"] = True
            logging.info(f"Set GitHub repo '{github_repo_name}' as origin.")
            # 5. Push branch
            last_dir = github_master_branch.strip().split("/")[-1]
            # Push the branch to GitHub
            if branch_name == last_dir:
                logging.info(f"Branch '{branch_name}' matches the master branch '{last_dir}'. Pushing directly to 'main'...")
                if not retry_subprocess(["git", "push", "-u", "origin", "master:main"], enable_proxy=True):
                    return 0
            else:
                logging.info(f"Branch '{branch_name}' does not match the master branch '{last_dir}'. Pushing to specific branch...")
                if not retry_subprocess(["git", "push", "-u", "origin", f"master:{branch_name}"], enable_proxy=True):
                    return 0
            success_flags["pushed_branch"] = True

            # 6. Push tags
            logging.info(f"Pushing tags for branch '{branch_name}'...")
            if not retry_subprocess(["git", "push", "--tags"], enable_proxy=True):
                return 0
            success_flags["pushed_tags"] = True

            # 7. Move the new log file to the date folder and rename it
            for file in os.listdir(git_tfs_log_path):
                if file.endswith(".txt"):
                    new_log_path = os.path.join(date_folder_path, f"{collection_name}_{project_name}_{branch_name}.log")
                    shutil.move(os.path.join(git_tfs_log_path, file), new_log_path)
                    logging.info(f"Moved log file '{file}' to '{new_log_path}'.")
                    success_flags["logs_moved"] = True

            # Only mark checkpoint if ALL steps passed
            if all(success_flags.values()):
                save_checkpoint(collection_name, project_name, branch_name)
                logging.info(f"Migration completed successfully for branch '{branch_name}'.")
            else:
                logging.error(f"Branch '{branch_name}' migration incomplete. Checklist: {success_flags}")
                return 0

        except Exception as e:
            logging.error(f"An error occurred while processing branch '{branch_name}': {e}")
            return 0

        finally:
            # Change back to the root directory
            os.chdir(root_dir)
            logging.info(f"Returned to root directory: {root_dir}")

    # Copy the TFVC_log folder to the script's root directory
    final_log_folder = os.path.join(root_dir, "TFVC_log")
    try:
        shutil.copytree(os.path.join(git_tfs_log_path, "TFVC_log"), final_log_folder, dirs_exist_ok=True)
        logging.info(f"Copied TFVC_log folder to script root directory: {final_log_folder}")
    except Exception as e:
        logging.warning(f"Failed to copy TFVC_log folder: {e}")
    logging.info("==================== TFVC Migration Finished ====================")
    return 1

def process_excel_file(tokens, row: dict):
    """Processes the Excel file and generates folder structures and config.json."""

    try:
        collection_name = str(row.get("COLLECTION_NAME", "") or "").strip()
        project_name = str(row.get("PROJECT_NAME", "") or "").replace("$/", "").strip()
        branch_list_raw = str(row.get("BRANCH_LIST", "") or "").strip()
        tfvc_master_branch = str(row.get("MASTER_BRANCH", "") or "").strip()
        github_organization = str(row.get("GITHUB_ORGANIZATION", "") or "").strip()
        github_repo_name = str(row.get("GITHUB_REPO", "") or "").strip()

        # Split branch_list by comma and strip spaces
        branch_list = [branch.strip() for branch in str(branch_list_raw).split(",") if branch.strip()]

        # Create folder structure
        collection_path = os.path.join(os.getcwd(), collection_name)
        project_path = os.path.join(collection_path, project_name)
        os.makedirs(project_path, exist_ok=True)

        # Decide where to place config.json
        if len(branch_list) == 1:
            branch_path = branch_list[0].replace("$/", "").strip().split("/")
            repo_path = os.path.join(project_path, *branch_path[1:]) if branch_path else project_path
            os.makedirs(repo_path, exist_ok=True)
            config_path = os.path.join(repo_path, "config.json")
        else:
            config_path = os.path.join(project_path, "config.json")

        # Generate JSON
        config_data = {
            "tfs_url": tokens.get("tfs_url", ""),
            "azure_devops_organization": collection_name,
            "azure_devops_project": project_name,
            "AZURE_DEVOPS_PAT": tokens.get("azure_devops_pat", ""),
            "github_token": tokens.get("github_token", ""),
            "github_organization": github_organization,
            "github_repo_name": github_repo_name,
            "specific_branches": branch_list,
            "tfvc_master_branch": tfvc_master_branch
        }

        # Write config.json
        with open(config_path, "w") as json_file:
            json.dump(config_data, json_file, indent=4)

        logging.info(
            f"Created config.json for: {collection_name} -> {project_name} "
            f"({'multiple branches' if len(branch_list) > 1 else branch_list[0]})"
        )

    except Exception as e:
        logging.error(f"Error processing row for {row.get('COLLECTION_NAME')} / {row.get('PROJECT_NAME')}: {e}")

def add_repo_to_runner_group(org, repo, github_token):
    try:
        headers = {"Authorization": f"token {github_token}", "Accept": "application/vnd.github+json"}
        response = requests.get(f"https://api.github.com/orgs/{org}/actions/runner-groups", headers=headers)
        runner_groups = response.json().get("runner_groups", [])
        group = next((g for g in runner_groups if g["name"] == "Migration-Runners"), None)
        if not group:
            raise Exception("Runner group 'Migration-Runners' not found")

        group_id = group["id"]
        repo_id = get_repo_id(org, repo, github_token)
        url = f"https://api.github.com/orgs/{org}/actions/runner-groups/{group_id}/repositories/{repo_id}"
        response = requests.put(url, headers=headers)
        if response.status_code not in [204, 201]:
            raise Exception(f"Failed to add repo to runner group: {response.status_code} - {response.text}")
        logging.info(f"Added {repo} to Migration-Runners runner group")
    except Exception as e:
        logging.error(f"Error adding repo to runner group: {e}")
        raise

def read_excel_rows(excel_file: str) -> list[dict]:
    """Reads the Excel input file and validates that both 'migration' and 'access' sheets exist. Returns rows only from the 'migration' sheet as a list of dictionaries."""
    try:
        # Load all sheet names
        xl = pd.ExcelFile(excel_file)
        required_sheets = ["migration", "access"]

        # Validate both sheets exist
        for sheet in required_sheets:
            if sheet not in xl.sheet_names:
                logging.error(f"Missing required sheet: {sheet}")
                return []

        # Load migration sheet
        df_migration = xl.parse("migration")

        # Drop empty rows
        df_migration = df_migration.dropna(how="all")

        # Convert to list of dicts
        rows = df_migration.to_dict(orient="records")

        logging.info(f"Loaded {len(rows)} rows from 'migration' sheet.")
        return rows

    except Exception as e:
        logging.error(f"Failed to read Excel file {excel_file}: {e}")
        return []

def main(excel_file: str):
    setup_logging()  # Setup logging at the start
    try:
        # Step 1: Validate Excel 
        if not os.path.isfile(excel_file):
            logging.error(f"Excel file not found at: {excel_file}")
            return
        git_tfs_log_path, date_folder_path = prompt_git_tfs_log_path()
        tokens = load_tokens()
        github_token = tokens.get("github_token", "")
        nexus_username = tokens.get("nexus_username", "")
        nexus_password = tokens.get("nexus_password", "")

        # Load all rows from Excel
        rows = read_excel_rows(excel_file)

        # Collect migration results
        migration_results = []

        for row in rows:
            collection_name = str(row.get("COLLECTION_NAME", "")or "").strip()
            project_name = str(row.get("PROJECT_NAME", "") or "").replace("$/", "").strip()
            checklist_passed = True
            comments = []

            # Initialize step flags
            json_created_flag = False
            repo_created_flag = False
            migration_done_flag = False
            access_assigned_flag = False
            secrets_added_flag = False
            row_result_recorded = False
            try:
                # Step 2: JSON creation for this row
                logging.info(f"[{collection_name}/{project_name}] Creating JSON...")
                process_excel_file(tokens, row)
                json_created_flag = True
            except Exception as e:
                checklist_passed = False
                comments.append(f"JSON creation failed: {e}")

            try:
                # Step 3: Repo creation
                if json_created_flag:
                    logging.info(f"[{collection_name}/{project_name}] Creating repository...")
                    create_repos_from_excel(row, github_token)
                    repo_created_flag = True
                else:
                    comments.append("Repo creation skipped due to JSON failure.")
            except Exception as e:
                checklist_passed = False
                comments.append(f"Repo creation failed: {e}")

            # Step 4: Migration (scan config.json inside the generated folder)
            try:
                if repo_created_flag:
                    found_config = False
                    # Build config.json path based on how process_excel_file creates it
                    branch_list_raw = str(row.get("BRANCH_LIST", "") or "").strip()
                    branch_list = [b.strip() for b in branch_list_raw.split(",") if b.strip()]

                    collection_path = os.path.join(os.getcwd(), collection_name)
                    project_path = os.path.join(collection_path, project_name)

                    if len(branch_list) > 1:
                        config_path = os.path.join(project_path, "config.json")
                    else:
                        branch_path_parts = branch_list[0].replace("$/", "").split("/")
                        repo_path = os.path.join(project_path, *branch_path_parts[1:]) if branch_path_parts else project_path
                        config_path = os.path.join(repo_path, "config.json")

                    if os.path.exists(config_path):
                        config = load_config(config_path)
                        if (config.get("azure_devops_project") == project_name and
                            config.get("azure_devops_organization") == collection_name):
                            found_config = True
                            result = clone_and_push_tfvc(config, git_tfs_log_path, date_folder_path)
                            if result == 0:  # fail → skip remaining steps for this repo
                                checklist_passed = False
                                comments.append("Migration failed during TFVC clone/push.")
                                # finalize failure result and move to next repo
                                migration_results.append({
                                    "COLLECTION_NAME": collection_name,
                                    "PROJECT_NAME": project_name,
                                    "BRANCH_LIST": row.get("BRANCH_LIST", ""),
                                    "STATUS": "FAILED",
                                    "COMMENTS": "; ".join(comments),
                                    "JSON_CREATED": json_created_flag,
                                    "REPO_CREATED": repo_created_flag,
                                    "MIGRATION_DONE": False,
                                    "ACCESS_ASSIGNED": False,
                                    "SECRETS_ADDED": False
                                })
                                row_result_recorded = True
                            else:
                                migration_done_flag = True

                    if not found_config and not migration_done_flag:
                        checklist_passed = False
                        comments.append("config.json not found for migration.")
                else:
                    comments.append("Migration skipped due to repo creation failure.")
            except Exception as e:
                checklist_passed = False
                comments.append(f"Migration failed: {e}")
                       
            # Step 5: Access assignment
            if migration_done_flag:
                try:
                    assign_repo_access_from_access_sheet(excel_file, row.get("GITHUB_ORGANIZATION", ""), row.get("GITHUB_REPO", ""), github_token)
                    access_assigned_flag = True
                except Exception as e:
                    comments.append(f"Access assignment failed: {e}")

                # Step 6: Secrets
                try:
                    process_excel_file_for_secrets(row, github_token, nexus_username, nexus_password) 
                    secrets_added_flag = True
                except Exception as e:
                    comments.append(f"Secrets failed: {e}")

                #get repo ID
                try:
                    repo_id = get_repo_id(row.get("GITHUB_ORGANIZATION", ""), row.get("GITHUB_REPO", ""), github_token)
                    update_excel_repo_id_column(excel_file, str(row.get("BRANCH_LIST", "") or ""), repo_id)
                except Exception as e:
                    comments.append(f"Failed to update Excel with repo ID: {e}")
                # Finalize status for this row
                checklist_passed = all([json_created_flag, repo_created_flag, migration_done_flag, access_assigned_flag, secrets_added_flag])
                migration_results.append({
                    "COLLECTION_NAME": collection_name,
                    "PROJECT_NAME": project_name,
                    "BRANCH_LIST": row.get("BRANCH_LIST", ""),
                    "STATUS": "SUCCESSFUL" if checklist_passed else "FAILED",
                    "COMMENTS": "; ".join(comments),
                    "JSON_CREATED": json_created_flag,
                    "REPO_CREATED": repo_created_flag,
                    "MIGRATION_DONE": migration_done_flag,
                    "ACCESS_ASSIGNED": access_assigned_flag,
                    "SECRETS_ADDED": secrets_added_flag
                })
                row_result_recorded = True
            # If we haven't recorded a result yet (e.g., JSON failed, repo failed, config.json missing),
            # record a failure outcome for this row.
            if not row_result_recorded:
                migration_results.append({
                    "COLLECTION_NAME": collection_name,
                    "PROJECT_NAME": project_name,
                    "BRANCH_LIST": row.get("BRANCH_LIST", ""),
                    "STATUS": "SUCCESSFUL" if checklist_passed else "FAILED",
                    "COMMENTS": "; ".join(comments),
                    "JSON_CREATED": json_created_flag,
                    "REPO_CREATED": repo_created_flag,
                    "MIGRATION_DONE": migration_done_flag,
                    "ACCESS_ASSIGNED": access_assigned_flag,
                    "SECRETS_ADDED": secrets_added_flag
                })

        # Step 7: Update Excel after all rows processed
        try:
            update_excel_with_status(excel_file, migration_results)
            logging.info("Excel updated with migration checklist and results.")
        except Exception as e:
            logging.error(f"Failed to update Excel with results: {e}")

    except Exception as e:
        logging.error(f"An error occurred in the migration script: {e}")

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="TFVC to GitHub migration orchestrator")
    parser.add_argument("--excel", required=True, help="Path to the Excel input file (e.g., /path/to/TFVC_Input_sheet.xlsx)",)
    args = parser.parse_args()
    main(args.excel)
