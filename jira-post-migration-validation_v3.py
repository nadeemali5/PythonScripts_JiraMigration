import requests
import logging
import argparse
from requests.auth import HTTPBasicAuth
from datetime import datetime
import sys
import os
import json
import csv
import time
from openpyxl import Workbook

def setup_logging(script_name, jira_project_key):
    if not os.path.exists('Logs'):
        os.makedirs('Logs')
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    log_filename = f"Logs/{script_name}_{jira_project_key}_{timestamp}_exception_log.log"
    logging.basicConfig(
        filename=log_filename,
        level=logging.INFO,
        format="%(asctime)s %(levelname)s %(message)s"
    )
    logging.info("Logging initialized.")

def read_jira_token(token_file_path):
    try:
        with open(token_file_path, 'r') as file:
            token = file.read().strip()
        logging.info("Successfully read JIRA token.")
        return token
    except Exception as e:
        logging.error(f"Failed to read JIRA token: {e}")
        sys.exit(1)

def get_jira_fields(session, jira_url, username, pat_token):
    url = f"{jira_url}/rest/api/3/field"
    headers = {'Content-Type': 'application/json'}
    try:
        response = session.get(url, headers=headers, auth=HTTPBasicAuth(username, pat_token), timeout=20)
        response.raise_for_status()
        fields = response.json()
        custom_field_mapping = {}
        for field in fields:
            if field["name"] == "TFS_WIT_ID":
                custom_field_mapping["TFS_WIT_ID"] = field["id"]
                logging.info(f"Found custom field 'TFS_WIT_ID': {field['id']}")
        return custom_field_mapping
    except requests.exceptions.RequestException as e:
        logging.error(f"Failed to get JIRA fields: {e}")
        return {}

def get_sprints(session, jira_url, board_id, username, pat_token):
    url = f"{jira_url}/rest/agile/1.0/board/{board_id}/sprint"
    headers = {'Content-Type': 'application/json'}
    startAt = 0
    maxResults = 50
    sprint_names = []
    try:
        while True:
            params = {"startAt": startAt, "maxResults": maxResults}
            response = session.get(url, headers=headers, auth=HTTPBasicAuth(username, pat_token), params=params, timeout=20)
            if 'Retry-After' in response.headers:
                wait_time = int(response.headers['Retry-After']) + 10
                logging.warning(f"Rate limit hit. Retrying after {wait_time} seconds.")
                time.sleep(wait_time)
                continue
            response.raise_for_status()
            sprints = response.json().get('values', [])
            if not sprints:
                break
            for sprint in sprints:
                sprint_names.append(sprint['name'])
                logging.info(f"Fetched sprint: {sprint['name']}")
            startAt += maxResults
        logging.info(f"Total Jira Sprints Fetched: {len(sprint_names)}")
        return sprint_names
    except requests.exceptions.RequestException as e:
        logging.error(f"Failed to fetch Jira sprints: {e}")
        return []

def process_tfs_json(json_file_path):
    def extract_items(obj, flat_data):
        if isinstance(obj, dict):
            if all(k in obj for k in ("comments", "attachments", "wit_links")):
                return
            for k, v in obj.items():
                if isinstance(v, dict) and all(x in v for x in ('comments', 'attachments', 'wit_links')):
                    flat_data[str(k)] = {
                        "TFS_Comments": len(v.get("comments", [])),
                        "TFS_Attachments": len(v.get("attachments", [])),
                        "TFS_Links": len(v.get("wit_links", []))
                    }
                else:
                    extract_items(v, flat_data)
        elif isinstance(obj, list):
            for item in obj:
                extract_items(item, flat_data)

    try:
        with open(json_file_path, 'r', encoding='utf-8-sig') as file:
            tfs_data = json.load(file)
        flat_data = {}
        extract_items(tfs_data, flat_data)
        logging.info(f"Processed TFS JSON. Work items loaded: {len(flat_data)}")
        return flat_data
    except Exception as e:
        logging.error(f"Failed to process TFS JSON: {e}")
        return {}

def read_tfs_sprint_names(csv_file_path):
    try:
        tfs_sprint_names = []
        with open(csv_file_path, 'r', encoding='utf-8-sig') as file:
            reader = csv.DictReader(file)
            for row in reader:
                name = row.get('jira_sprint_name', '').strip()
                if name:
                    tfs_sprint_names.append(name)
        logging.info(f"Read {len(tfs_sprint_names)} TFS sprint names from CSV.")
        return tfs_sprint_names
    except Exception as e:
        logging.error(f"Failed to read TFS sprint names: {e}")
        return []

def write_sprint_comparison_to_sheet(workbook, tfs_sprint_names, jira_sprint_names):
    sheet2 = workbook.create_sheet(title="Sprint Comparison")
    header2 = ["TFS Sprint Name", "JIRA Sprint Name", "Match"]
    sheet2.append(header2)
    for tfs in tfs_sprint_names:
        match = tfs in jira_sprint_names
        matched_jira = tfs if match else ""
        sheet2.append([tfs, matched_jira, match])
    for jira in jira_sprint_names:
        if jira not in tfs_sprint_names:
            sheet2.append(["", jira, False])
    logging.info(f"Completed writing sprint comparison with {len(tfs_sprint_names)} entries.")

def write_issues_to_excel(project_issues, tfs_summary, script_name, project_key, tfs_sprint_names, jira_sprint_names):
    if not project_issues and not tfs_summary:
        logging.warning("No issues or TFS data found to write to Excel.")
        print("No issues or TFS data found to write to Excel.")
        return

    workbook = Workbook()
    sheet1 = workbook.active
    sheet1.title = "Jira Issues and TFS Data"
    header = [
        "TFS_JSON_ID", "TFS_Attachments", "TFS_Comments", "TFS_Links",
        "JIRA_Issue_Key", "JIRA_Attachments", "JIRA_Comments", "JIRA_TFS_WIT_ID", "JIRA_Links",
        "Match", "Status"
    ]
    sheet1.append(header)
    mismatch_wb = Workbook()
    mismatch_ws = mismatch_wb.active
    mismatch_ws.title = "Mismatch Work Items"
    mismatch_ws.append(header)

    total_tfs_items = len(tfs_summary)
    total_jira_items = len(project_issues)
    tfs_ids = set(tfs_summary.keys())
    jira_ids = set(issue.get("JIRA_TFS_WIT_ID") for issue in project_issues)
    compared_ids = sorted(jira_ids)
    not_compared_tfs = sorted(tfs_ids - jira_ids)
    not_compared_jira = sorted(jira_ids - tfs_ids)

    # Sprint comparison stats
    tfs_sprints_set = set(tfs_sprint_names)
    jira_sprints_set = set(jira_sprint_names)
    compared_sprints = sorted(tfs_sprints_set & jira_sprints_set)
    tfs_sprints_not_in_jira = sorted(tfs_sprints_set - jira_sprints_set)
    jira_sprints_not_in_tfs = sorted(jira_sprints_set - tfs_sprints_set)

    total = 0
    success = 0
    failed = 0

    tfs_id_map = {wit_id: data for wit_id, data in tfs_summary.items()}
    jira_id_map = {issue.get("JIRA_TFS_WIT_ID"): issue for issue in project_issues}
    all_ids = tfs_ids | jira_ids

    for wit_id in sorted(all_ids, key=lambda x: (x is None, x)):
        tfs_data = tfs_id_map.get(wit_id, {
            "TFS_Comments": "",
            "TFS_Attachments": "",
            "TFS_Links": ""
        })
        issue = jira_id_map.get(wit_id, {
            "JIRA_Issue_Key": "",
            "JIRA_Attachments": "",
            "JIRA_Comments": "",
            "JIRA_TFS_WIT_ID": wit_id,
            "JIRA_Links": ""
        })
        total += 1
        try:
            if (tfs_data["TFS_Comments"] != "" and tfs_data["TFS_Attachments"] != "" and tfs_data["TFS_Links"] != "" and
                issue["JIRA_Comments"] != "" and issue["JIRA_Attachments"] != "" and issue["JIRA_Links"] != ""):
                match = (
                    int(issue["JIRA_Comments"]) == int(tfs_data["TFS_Comments"]) and
                    int(issue["JIRA_Attachments"]) == int(tfs_data["TFS_Attachments"]) and
                    int(issue["JIRA_Links"]) == int(tfs_data["TFS_Links"])
                )
            else:
                match = False
            status = "Successful" if match else "Failed"
            if match:
                success += 1
            else:
                failed += 1
                mismatch_ws.append([
                    wit_id, tfs_data["TFS_Attachments"], tfs_data["TFS_Comments"], tfs_data["TFS_Links"],
                    issue["JIRA_Issue_Key"], issue["JIRA_Attachments"], issue["JIRA_Comments"],
                    issue["JIRA_TFS_WIT_ID"], issue["JIRA_Links"], match, status
                ])
        except Exception as e:
            match = False
            status = "Failed"
            failed += 1
            logging.error(f"Error processing WIT ID {wit_id}: {e}")

        sheet1.append([
            wit_id, tfs_data["TFS_Attachments"], tfs_data["TFS_Comments"], tfs_data["TFS_Links"],
            issue["JIRA_Issue_Key"], issue["JIRA_Attachments"], issue["JIRA_Comments"],
            issue["JIRA_TFS_WIT_ID"], issue["JIRA_Links"], match, status
        ])

    # Sprint comparison
    write_sprint_comparison_to_sheet(workbook, tfs_sprint_names, jira_sprint_names)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_file = f"Post_Migration_excel_{project_key}_{timestamp}.xlsx"
    mismatch_file = f"Mismatch_WorkItems_{project_key}_{timestamp}.xlsx"
    workbook.save(output_file)
    mismatch_wb.save(mismatch_file)

    summary_msg = (
        f"\nSummary:\n"
        f"Total TFS Work Items (from JSON): {total_tfs_items}\n"
        f"Total JIRA Work Items (from API): {total_jira_items}\n"
        f"Total Compared Work Items (JIRA TFS_WIT_IDs): {len(compared_ids)}\n"
        f"Successful: {success}\n"
        f"Failed: {failed}\n"
        f"Work Items present in TFS but not compared (missing in JIRA): {len(not_compared_tfs)}\n"
        f"  IDs: {', '.join(not_compared_tfs) if not_compared_tfs else 'None'}\n"
        f"Work Items present in JIRA but not in TFS: {len(not_compared_jira)}\n"
        f"  IDs: {', '.join(not_compared_jira) if not_compared_jira else 'None'}\n"
        f"--- Sprint Comparison ---\n"
        f"Total TFS Sprints: {len(tfs_sprint_names)}\n"
        f"Total JIRA Sprints: {len(jira_sprint_names)}\n"
        f"Total Compared Sprints: {len(compared_sprints)}\n"
        f"TFS Sprints Not in JIRA: {len(tfs_sprints_not_in_jira)}\n"
        f"  Names: {', '.join(tfs_sprints_not_in_jira) if tfs_sprints_not_in_jira else 'None'}\n"
        f"JIRA Sprints Not in TFS: {len(jira_sprints_not_in_tfs)}\n"
        f"  Names: {', '.join(jira_sprints_not_in_tfs) if jira_sprints_not_in_tfs else 'None'}\n"
        f"Mismatch File: {mismatch_file}"
    )
    logging.info(summary_msg)
    print(summary_msg)
    logging.info("Script execution completed.")

def get_project_issues(session, jira_url, project_key, custom_field_ids, username, pat_token, total=None):
    url = f"{jira_url}/rest/api/3/search/jql"
    headers = {'Content-Type': 'application/json', 'Accept': 'application/json'}
    timeout_sec = 30
    project_issues = []
    retry_delay = 10
    next_page_token = None

    while True:
        body = {
            "jql": f'project = "{project_key}"',
            "fields": ["attachment", "comment", "issuelinks", custom_field_ids["TFS_WIT_ID"]],
        }
        if next_page_token:
            body["nextPageToken"] = next_page_token

        try:
            response = session.post(
                url,
                headers=headers,
                auth=HTTPBasicAuth(username, pat_token),
                json=body,
                timeout=timeout_sec
            )

            if response.status_code == 429:
                retry_after = int(response.headers.get('Retry-After', retry_delay))
                logging.warning(f"[Rate Limit] Hit 429. Retrying in {retry_after + 5} seconds.")
                time.sleep(retry_after + 5)
                continue

            response.raise_for_status()
            data = response.json()
            issues = data.get('issues', [])
            if total is None:
                total = data.get('total') or data.get('maxResults')

            if not issues:
                break

            for issue in issues:
                try:
                    tfs_wit_id = issue.get("fields", {}).get(custom_field_ids["TFS_WIT_ID"], "")
                    if isinstance(tfs_wit_id, float):
                        tfs_wit_id = str(int(tfs_wit_id))
                    else:
                        tfs_wit_id = str(tfs_wit_id)

                    project_issues.append({
                        "JIRA_Issue_Key": issue.get("key", ""),
                        "JIRA_Attachments": len(issue.get("fields", {}).get("attachment", [])),
                        "JIRA_Comments": issue.get("fields", {}).get("comment", {}).get("total", ""),
                        "JIRA_TFS_WIT_ID": tfs_wit_id,
                        "JIRA_Links": len(issue.get("fields", {}).get("issuelinks", []))
                    })
                except Exception as e:
                    logging.error(f"[Issue Parse Error] {issue.get('key', 'Unknown')}: {e}")
                    project_issues.append({
                        "JIRA_Issue_Key": issue.get("key", ""),
                        "JIRA_Attachments": "",
                        "JIRA_Comments": "",
                        "JIRA_TFS_WIT_ID": "",
                        "JIRA_Links": ""
                    })

            next_page_token = data.get("nextPageToken")
            if not next_page_token:
                break

        except requests.exceptions.Timeout:
            logging.error(f"[Timeout] Fetching issues from index. Skipping this batch.")
            continue
        except requests.exceptions.RequestException as req_err:
            try:
                error_detail = response.json()
            except Exception:
                error_detail = response.text
            logging.error(f"[Request Exception] While fetching issues: {req_err} | Details: {error_detail}")
            break

    logging.info(f"Total JIRA issues fetched: {len(project_issues)}")
    return project_issues

def main():
    parser = argparse.ArgumentParser(description='Validate JIRA issues against TFS export data.')
    parser.add_argument('--jira_url', required=True, help='JIRA base URL')
    parser.add_argument('--jira_project_key', required=True, help='JIRA Project Key')
    parser.add_argument('--username', required=True, help='JIRA Username')
    parser.add_argument('--jira_token_file_path', required=True, help='Path to JIRA token file')
    parser.add_argument('--tfs_json_file_path', required=True, help='Path to TFS JSON file')
    parser.add_argument('--tfs_csv_file_path', required=True, help='Path to TFS Sprint CSV file')
    parser.add_argument('--board_id', required=True, help='JIRA Board ID')
    args = parser.parse_args()
    script_name = os.path.basename(__file__).replace('.py', '')
    setup_logging(script_name, args.jira_project_key)
    logging.info("Script execution started.")
    print("Script execution started.")

    pat_token = read_jira_token(args.jira_token_file_path)
    with requests.Session() as session:
        custom_field_ids = get_jira_fields(session, args.jira_url, args.username, pat_token)
        if not custom_field_ids:
            logging.error("No custom field ID found. Please verify JIRA configuration.")
            print("No custom field ID found. Exiting.")
            sys.exit(1)
        jira_sprint_names = get_sprints(session, args.jira_url, args.board_id, args.username, pat_token)
        project_issues = get_project_issues(session, args.jira_url, args.jira_project_key, custom_field_ids, args.username, pat_token)
        tfs_summary = process_tfs_json(args.tfs_json_file_path)
        tfs_sprint_names = read_tfs_sprint_names(args.tfs_csv_file_path)
        write_issues_to_excel(project_issues, tfs_summary, script_name, args.jira_project_key, tfs_sprint_names, jira_sprint_names)

if __name__ == "__main__":
    main()